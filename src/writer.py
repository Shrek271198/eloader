def clear_output():
    import os, re, os.path

    mypath = "output"
    for root, dirs, files in os.walk(mypath):
        for file in files:
            os.remove(os.path.join(root, file))


def mcc_writer(els, STANDARD):
    """This method accepts an els object and using the MCC
    template produces a summary for each MCC.

    It does this by first making a copy of the MCC Template file.

    Then making a modification of the `Template for MCC' sheet and
    appending the appropriate detail.

    Finally, it removes the template sheet and saves the file.
    """

    from openpyxl import load_workbook
    from shutil import copyfile
    from sys import exit, exc_info
    from copy import copy
    from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties

    mcc_template = "src/templates/mcc_template.xlsx"

    for mcc in els.mccl:

        # Copy and Create MCC excel files
        mcc_title = "ELECTRICAL LOADS LIST SUBSTATION - {}".format(mcc.name)
        mcc_output_file = "output/{}.xlsx".format(mcc_title)
        try:
            copyfile(mcc_template, mcc_output_file)
        except IOError as e:
            print("Unable to copy file. %s" % e)
            exit(1)
        except:
            print("Unexpected error:", exc_info())
            exit(1)

        wb = load_workbook(filename=mcc_output_file)
        ws = wb["Template for MCC"]

        # Add Project Title
        ws["G1"] = STANDARD.project_name
        ws["G4"] = mcc_title

        # Add Project Details
        ws["R1"] = STANDARD.project
        ws["R2"] = STANDARD.revision
        ws["R3"] = STANDARD.prepared_by
        ws["R4"] = STANDARD.date_prepared
        ws["R5"] = STANDARD.approved_by
        ws["R6"] = STANDARD.date_approved

        # Update Inserted Row style
        count =  1
        for me in mcc.mel:
            # Insert empty rows
            ws.insert_rows(9)
            style_cell = "B" + str(9+count)
            # Apply style
            for row in ws.iter_cols(min_row=9, max_row=9, min_col=1, max_col=19):
                for cell in row:
                    cell.style = copy(ws[style_cell].style)
                    cell.font = copy(ws[style_cell].font)
                    cell.border = copy(ws[style_cell].border)
                    cell.fill = copy(ws[style_cell].fill)
                    cell.number_format = copy(ws[style_cell].number_format)
                    cell.protection = copy(ws[style_cell].protection)
                    cell.alignment = copy(ws[style_cell].alignment)

        # Insert Mechanical Equipment data
        start_row = 9
        for i, me in enumerate(mcc.mel):
            ws.cell(row=start_row+i, column=1).value = me.tag_number
            ws.cell(row=start_row+i, column=2).value = me.rev
            ws.cell(row=start_row+i, column=3).value = me.area
            ws.cell(row=start_row+i, column=4).value = me.type
            ws.cell(row=start_row+i, column=5).value = me.starter_type
            ws.cell(row=start_row+i, column=6).value = me.voltage
            ws.cell(row=start_row+i, column=7).value = me.operation_mode
            ws.cell(row=start_row+i, column=8).value = me.name
            ws.cell(row=start_row+i, column=9).value = me.workpack
            ws.cell(row=start_row+i, column=10).value = me.installed_kw
            ws.cell(row=start_row+i, column=11).value = me.power_factor
            ws.cell(row=start_row+i, column=12).value = me.kva
            ws.cell(row=start_row+i, column=13).value = me.load_factor
            ws.cell(row=start_row+i, column=14).value = me.diversity_utilisation
            ws.cell(row=start_row+i, column=15).value = me.avg_load_factor
            ws.cell(row=start_row+i, column=16).value = me.max_kw
            ws.cell(row=start_row+i, column=17).value = me.max_kvar
            ws.cell(row=start_row+i, column=18).value = me.max_kva
            ws.cell(row=start_row+i, column=19).value = me.avg_load_kw

        # Insert Misc Equiptment data
        start_row = 9 + len(mcc.mel) + 2
        for i, misc in enumerate([mcc.lighting, mcc.ups, mcc.field_equipment]):
            ws.cell(row=start_row+i, column=6).value = 240
            ws.cell(row=start_row+i, column=10).value = misc.installed_kw
            ws.cell(row=start_row+i, column=11).value = misc.power_factor
            ws.cell(row=start_row+i, column=12).value = misc.kva
            ws.cell(row=start_row+i, column=13).value = misc.load_factor
            ws.cell(row=start_row+i, column=14).value = misc.diversity_utilisation
            ws.cell(row=start_row+i, column=15).value = misc.avg_load_factor
            ws.cell(row=start_row+i, column=16).value = misc.max_kw
            ws.cell(row=start_row+i, column=17).value = misc.max_kvar
            ws.cell(row=start_row+i, column=18).value = misc.max_kva
            ws.cell(row=start_row+i, column=19).value = misc.avg_load_kw

        # Insert Totals
        start_row = 9 + len(mcc.mel) + 6
        ws.cell(row=start_row, column=10).value = mcc.total_installed_kw
        ws.cell(row=start_row, column=12).value = mcc.total_kva
        ws.cell(row=start_row, column=16).value = mcc.total_max_kw
        ws.cell(row=start_row, column=17).value = mcc.total_max_kvar
        ws.cell(row=start_row, column=18).value = mcc.total_max_kva
        ws.cell(row=start_row, column=19).value = mcc.total_avg_load_kw

        # Insert Contingency data
        start_row = 9 + len(mcc.mel) + 8
        ws.cell(row=start_row, column=10).value = str(int(mcc.contingency_factor_percent * 100))+'%'
        ws.cell(row=start_row+1, column=10).value = mcc.spare_starters
        ws.cell(row=start_row+2, column=10).value = mcc.contingency_load
        ws.cell(row=start_row+3, column=10).value = mcc.total_spare_allocation
        ws.cell(row=start_row+4, column=10).value = mcc.total_mcc_load_allowed

        # Adjust print area
        ws.print_area = "A1:S{}".format(21+len(mcc.mel))
        # Rename sheet
        ws.title = mcc.name
        # Save file
        wb.save(filename=mcc_output_file)


def els_writer(els, STANDARD):
    """This method accepts an els object and using the MCC
    template produces a summary for each MCC.

    It does this by first making a copy of the MCC Template file.

    Then making a modification of the `Electrical Load Summary Template' sheet and
    appending the appropriate detail.

    Finally, it removes the template sheet and saves the file.
    """

    from openpyxl import load_workbook
    from shutil import copyfile
    from sys import exit, exc_info
    from copy import copy
    from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties

    els_template = "src/templates/electrical_load_summary_template.xlsx"

    # Copy and Create excel file
    power_summary_output_file = "output/POWER SUMMARY.xlsx"
    try:
        copyfile(els_template, power_summary_output_file)
    except IOError as e:
        print("Unable to copy file. %s" % e)
        exit(1)
    except:
        print("Unexpected error:", exc_info())
        exit(1)

    wb = load_workbook(filename=power_summary_output_file)
    ws = wb["ELECTRICAL LOAD SUMMARY"]

    # Add Project Title
    ws["D1"] = STANDARD.project_name
    ws["D4"] = "MCC LOAD DISTRIBUTION AND TX SIZING"

    # Add Project Details
    ws["L1"] = STANDARD.project
    ws["L2"] = STANDARD.revision
    ws["L3"] = STANDARD.prepared_by
    ws["L4"] = STANDARD.date_prepared
    ws["L5"] = STANDARD.approved_by
    ws["L6"] = STANDARD.date_approved

    # Update Inserted Row style
    count =  1
    for mcc in els.mccl:
        # Insert empty rows
        offset = 10
        style_cell = "A2"
        ws.insert_rows(offset)
        # Apply style
        for row in ws.iter_cols(min_row=offset, max_row=offset, min_col=1, max_col=13):
            for cell in row:
                cell.style = copy(ws[style_cell].style)
                cell.font = copy(ws[style_cell].font)
                cell.border = copy(ws[style_cell].border)
                cell.fill = copy(ws[style_cell].fill)
                cell.number_format = copy(ws[style_cell].number_format)
                cell.protection = copy(ws[style_cell].protection)
                cell.alignment = copy(ws[style_cell].alignment)

    # Insert MCC data
    start_row = offset
    for i, mcc in enumerate(els.mccl):
        ws.cell(row=start_row+i, column=1).value = mcc.name
        ws.cell(row=start_row+i, column=2).value = mcc.max_voltage
        #ws.cell(row=start_row+i, column=3).value = ""
        ws.cell(row=start_row+i, column=4).value = mcc.total_installed_kw
        ws.cell(row=start_row+i, column=5).value = mcc.total_kva
        ws.cell(row=start_row+i, column=6).value = mcc.total_max_kw
        ws.cell(row=start_row+i, column=7).value = mcc.total_max_kvar
        ws.cell(row=start_row+i, column=8).value = mcc.total_max_kva
        ws.cell(row=start_row+i, column=9).value = mcc.total_avg_load_kva
        ws.cell(row=start_row+i, column=10).value = mcc.contingency_factor
        ws.cell(row=start_row+i, column=11).value = mcc.total_actual_contingency
        ws.cell(row=start_row+i, column=12).value = mcc.tx_size
        ws.cell(row=start_row+i, column=13).value = mcc.spare_tx

    # Insert Network Losses
    start_row = offset + len(els.mccl) + 1
    ws.cell(row=start_row, column=6).value = els.network_loss_kw
    ws.cell(row=start_row, column=7).value = els.network_loss_kvar
    ws.cell(row=start_row, column=8).value = els.network_loss_kva

    # Insert Totals
    start_row = offset + len(els.mccl) + 2
    ws.cell(row=start_row, column=4).value = els.connected_load_kw
    ws.cell(row=start_row, column=5).value = els.connected_load_kva
    ws.cell(row=start_row, column=6).value = els.max_demand_kw
    ws.cell(row=start_row, column=7).value = els.max_demand_kvar
    ws.cell(row=start_row, column=8).value = els.max_demand_kva
    ws.cell(row=start_row, column=9).value = els.ave_load_kva
    ws.cell(row=start_row, column=10).value = els.contingency_factor_kva
    ws.cell(row=start_row, column=11).value = els.total_actual_contingency

    # Adjust print area
    ws.print_area = "A1:M{}".format(59+len(els.mccl))
    # Save file
    wb.save(filename=power_summary_output_file)
