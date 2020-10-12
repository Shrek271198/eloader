from datetime import date
from enum import Enum
from math import acos, tan, ceil
from typing import List

from data import (CONTINGENCY_TABLE, LOAD_FACTOR, MOTOR_POWER_FACTOR,
                  VSD_CONTINGENCY)
from dataclasses import dataclass, field
from utility import round_up, vlookup


class OperationMode(Enum):
    DUTY = 1
    STANDBY = 2


@dataclass
class ProjectStandard:
    """ Class for storing Project standard details."""

    field_isolator: str
    project_name: str
    folder_name: str
    folder_path: str

    mcc_to_vsd: int
    mcc_to_fi: int
    vsd_to_fi: int
    vsd_mcc_to_motor: int

    project: str
    revision: str
    prepared_by: str
    date_prepared: date
    approved_by: str
    date_approved: date

    avg_count: float
    starters: float

    lighting_load: float
    ups_load: float
    fe_dist_load: float

    min_thermistor: float
    max_thermistor: float

    min_rtd: float
    max_rtd: float

    substation_cost: float
    max_cable_size: float


@dataclass
class LightingEquipment:
    """Class for Lighting Equipment details."""

    installed_kw: float

    power_factor: float = 1
    load_factor: float = 0.75
    diversity_utilisation: float = 0.9
    efficiency: float = 1

    kva: float = field(init=False)
    avg_load_factor: float = field(init=False)
    max_kw: float = field(init=False)
    max_kvar: float = field(init=False)
    max_kva: float = field(init=False)
    avg_load_kw: float = field(init=False)
    avg_load_kva: float = field(init=False)

    def __post_init__(self):

        try:
            self.kva = round_up(
                self.installed_kw / self.efficiency / self.power_factor, 1
            )
        except:
            self.kva = 0

        self.avg_load_factor = round_up(
            self.load_factor * self.diversity_utilisation, 3
        )

        try:
            self.max_kw = round_up(self.installed_kw * self.load_factor, 1)
        except:
            self.max_kw = 0

        self.max_kvar = self.max_kw

        self.max_kva = round_up(self.kva * self.load_factor, 1)

        try:
            self.avg_load_kw = round_up(self.installed_kw * self.avg_load_factor, 1)
        except:
            self.avg_load_kw = 0

        try:
            self.avg_load_kva = round_up(self.kva * self.avg_load_factor, 1)
        except:
            self.avg_load_kva = 0


@dataclass
class UPSEquipment:
    """Class for UPS Equipment details."""

    installed_kw: float

    power_factor: float = 1
    load_factor: float = 0.75
    diversity_utilisation: float = 0.9
    efficiency: float = 1

    kva: float = field(init=False)
    avg_load_factor: float = field(init=False)
    max_kw: float = field(init=False)
    max_kvar: float = field(init=False)
    max_kva: float = field(init=False)
    avg_load_kw: float = field(init=False)
    avg_load_kva: float = field(init=False)

    def __post_init__(self):

        try:
            self.kva = round_up(
                self.installed_kw / self.efficiency / self.power_factor, 1
            )
        except:
            self.kva = 0

        self.avg_load_factor = round_up(
            self.load_factor * self.diversity_utilisation, 3
        )

        try:
            self.max_kw = round_up(self.installed_kw * self.load_factor, 1)
        except:
            self.max_kw = 0

        self.max_kvar = self.max_kw

        self.max_kva = round_up(self.kva * self.load_factor, 1)

        try:
            self.avg_load_kw = round(self.installed_kw * self.avg_load_factor, 1)
        except:
            self.avg_load_kw = 0

        try:
            self.avg_load_kva = round(self.kva * self.avg_load_factor, 1)
        except:
            self.avg_load_kva = 0

@dataclass
class FieldEquipment:
    """Class for Field Equipment details."""

    installed_kw: float

    power_factor: float = 1
    load_factor: float = 0.75
    diversity_utilisation: float = 0.9
    efficiency: float = 1

    kva: float = field(init=False)
    avg_load_factor: float = field(init=False)
    max_kw: float = field(init=False)
    max_kvar: float = field(init=False)
    max_kva: float = field(init=False)
    avg_load_kw: float = field(init=False)
    avg_load_kva: float = field(init=False)

    def __post_init__(self):

        try:
            self.kva = round_up(
                self.installed_kw / self.efficiency / self.power_factor, 1
            )
        except:
            self.kva = 0

        self.avg_load_factor = round_up(
            self.load_factor * self.diversity_utilisation, 3
        )

        try:
            self.max_kw = round_up(self.installed_kw * self.load_factor, 1)
        except:
            self.max_kw = 0

        self.max_kvar = self.max_kw

        self.max_kva = round_up(self.kva * self.load_factor, 1)

        try:
            self.avg_load_kw = round_up(self.installed_kw * self.avg_load_factor, 1)
        except:
            self.avg_load_kw = 0

        try:
            self.avg_load_kva = round_up(self.kva * self.avg_load_factor, 1)
        except:
            self.avg_load_kva = 0


@dataclass
class MechanicalEquipment:
    """Class for Mechanical Equipment details."""

    area: str
    type: str
    number: str
    name: str
    workpack: str
    mcc_number: str
    installed_kw: float
    starter_type: str
    voltage: float
    operation_mode: OperationMode
    rev: str
    procurement_rating: str
    tag_number: str = field(init=False)
    contingency_factor: float = field(init=False)

    power_factor: float = field(init=False)
    efficiency: float = field(init=False)
    kva: float = field(init=False)
    load_factor: float = field(init=False)
    diversity_utilisation: float = field(init=False)
    avg_load_factor: float = field(init=False)
    max_kw: float = field(init=False)
    max_kvar: float = field(init=False)
    max_kva: float = field(init=False)
    avg_load_kw: float = field(init=False)
    avg_load_kva: float = field(init=False)
    spare_capacity: float = field(init=False)

    def __post_init__(self):

        self.tag_number = self.area + self.type + self.number

        self.efficiency = vlookup(self.installed_kw, MOTOR_POWER_FACTOR, 1)

        if self.starter_type in ["VSD", "VSD Dual"]:
            self.power_factor = 0.9
        else:
            self.power_factor = vlookup(self.installed_kw, MOTOR_POWER_FACTOR, 2)

        self.kva = round(self.installed_kw / self.efficiency / self.power_factor, 1)

        if self.operation_mode == 2:
            self.load_factor = 0
        else:
            self.load_factor = vlookup(self.type, LOAD_FACTOR, 2)

        if self.operation_mode == 2:
            self.diversity_utilisation = 0
        else:
            self.diversity_utilisation = vlookup(self.type, LOAD_FACTOR, 3)

        self.avg_load_factor = round_up(
            self.load_factor * self.diversity_utilisation, 3
        )

        self.max_kw = round_up(self.installed_kw * self.load_factor, 1)

        self.max_kvar = round(self.max_kw * round(tan(acos(self.power_factor)), 2), 1)

        self.max_kva = round(self.kva * self.load_factor, 1)

        self.avg_load_kw = round(self.installed_kw * self.avg_load_factor, 1)

        self.avg_load_kva = round(self.kva * self.avg_load_factor, 1)

        self.contingency_factor = vlookup(self.procurement_rating, CONTINGENCY_TABLE, 1)

        self.spare_capacity = round(self.contingency_factor * round((self.kva*self.avg_load_factor), 1), 2)


@dataclass
class MotorControlCenter:
    """Class for storing the electrical load details of a MCC."""

    name: str
    lighting: LightingEquipment
    ups: UPSEquipment
    field_equipment: FieldEquipment
    mel: List[MechanicalEquipment] = field(default_factory=list)

    total_installed_kw: float = field(init=False)
    total_kva: float = field(init=False)
    total_max_kw: float = field(init=False)
    total_max_kvar: float = field(init=False)
    total_max_kva: float = field(init=False)
    total_avg_load_kw: float = field(init=False)
    total_avg_load_kva: float = field(init=False)

    contingency_factor_percent = float = 0.2
    misc_starters: int = 3
    spare_starters: int = field(init=False)
    contingency_load: float = field(init=False)
    total_spare_allocation: float = field(init=False)
    total_mcc_load_allowed: float = field(init=False)

    max_voltage: float = field(init=False)
    contingency_factor: float = field(init=False)
    total_actual_contingency: float = field(init=False)
    tx_size: float = field(init=False)
    spare_tx: int = field(init=False)


    def __post_init__(self):
        self.total_installed_kw = round(
            (
                sum(e.installed_kw for e in self.mel)
                + self.lighting.installed_kw
                + self.ups.installed_kw
                + self.field_equipment.installed_kw
            ),
            1,
        )

        self.total_kva = round(
            (
                sum(e.kva for e in self.mel)
                + self.lighting.kva
                + self.ups.kva
                + self.field_equipment.kva
            ),
            1,
        )

        self.total_max_kw = round(
            (
                sum(e.max_kw for e in self.mel)
                + self.lighting.max_kw
                + self.ups.max_kw
                + self.field_equipment.max_kw
            ),
            1,
        )

        self.total_max_kvar = round(
            (
                sum(e.max_kvar for e in self.mel)
                + self.lighting.max_kvar
                + self.ups.max_kvar
                + self.field_equipment.max_kvar
            ),
            1,
        )

        self.total_max_kva = round(
            (
                sum(e.max_kva for e in self.mel)
                + self.lighting.max_kva
                + self.ups.max_kva
                + self.field_equipment.max_kva
            ),
            1,
        )

        self.total_avg_load_kw = round(
            (
                sum(e.avg_load_kw for e in self.mel)
                + self.lighting.avg_load_kw
                + self.ups.avg_load_kw
                + self.field_equipment.avg_load_kw
            ),
            1,
        )

        self.total_avg_load_kva = round(
            (
                sum(e.avg_load_kva for e in self.mel)
                + self.lighting.avg_load_kva
                #+ self.ups.avg_load_kva
                + self.field_equipment.avg_load_kva
            ),
            1,
        )

        self.spare_starters = round_up(
            (len(self.mel) + self.misc_starters) * self.contingency_factor_percent
        )

        # Init contingency_load
        installed_kw = round(sum(me.installed_kw for me in self.mel), 2)
        max_kva = sum(me.max_kva for me in self.mel)
        avg_load_kw = sum(me.avg_load_kw for me in self.mel)
        avg_starter_load = round(installed_kw / len(self.mel), 2)
        contingency_spare_starters = sum(
            me.contingency_factor for me in self.mel
        ) / len(self.mel)
        self.contingency_load = min(
            [load for load in VSD_CONTINGENCY if load >= avg_starter_load]
        )

        self.total_spare_allocation = self.contingency_load * self.spare_starters

        self.total_mcc_load_allowed = (
            self.total_spare_allocation + self.total_installed_kw
        )

        self.max_voltage = max(me.voltage for me in self.mel)

        self.contingency_factor = round_up(sum(me.spare_capacity for me in self.mel), 0)

        self.total_actual_contingency = self.contingency_factor + self.total_max_kva

        if self.total_actual_contingency <= 375:
            self.tx_size = 500
        elif self.total_actual_contingency <= 562.5:
            self.tx_size = 750
        elif self.total_actual_contingency <= 750:
            self.tx_size = 1000
        elif self.total_actual_contingency <= 1125:
            self.tx_size = 1500
        elif self.total_actual_contingency <= 1500:
            self.tx_size = 2000
        elif self.total_actual_contingency <= 1875:
            self.tx_size = 2500
        elif self.total_actual_contingency <= 3750:
            self.tx_size = 5000
        elif self.total_actual_contingency <= 6000:
            self.tx_size = 8000
        elif self.total_actual_contingency <= 7500:
            self.tx_size = 10000
        elif self.total_actual_contingency <= 8250:
            self.tx_size = 11000

        self.spare_tx = int(((self.tx_size - self.total_actual_contingency)/self.tx_size)*100)


@dataclass
class ElectricalLoadSummary:
    """Class for storing the MCC summary data."""


    mccl: List[MotorControlCenter] = field(default_factory=list)

    connected_load_kw: int = field(init=False)
    connected_load_kva: int = field(init=False)
    max_demand_kw: int = field(init=False)
    max_demand_kvar: int = field(init=False)
    max_demand_kva: int = field(init=False)
    ave_load_kva: int = field(init=False)
    contingency_factor_kva: int = field(init=False)
    total_actual_contingency: int = field(init=False)
    tx_size: int = field(init=False)
    spare_tx: int = field(init=False)

    network_loss_kw: int = field(init=False)
    network_loss_kvar: int = field(init=False)
    network_loss_kva: int = field(init=False)


    def __post_init__(self):

        self.connected_load_kw = round_up(sum(mcc.total_installed_kw for mcc in self.mccl))

        self.connected_load_kva = round_up(sum(mcc.total_kva for mcc in self.mccl))

        self.network_loss_kw = int(0.02 * (sum(ceil(mcc.total_max_kw) for mcc in self.mccl)))

        self.max_demand_kw = int(sum(mcc.total_max_kw for mcc in self.mccl) + self.network_loss_kw)

        self.network_loss_kvar = int(0.02 * (sum(ceil(mcc.total_max_kvar) for mcc in self.mccl)))

        self.max_demand_kvar = int(sum(ceil(mcc.total_max_kvar) for mcc in self.mccl) + self.network_loss_kvar)

        self.network_loss_kva = ceil(0.02 * (sum(mcc.total_max_kva for mcc in self.mccl)))

        self.max_demand_kva = int(sum(mcc.total_max_kva for mcc in self.mccl) + self.network_loss_kva)

        self.ave_load_kva = int(sum(mcc.total_avg_load_kva for mcc in self.mccl))

        self.contingency_factor_kva = int(sum(mcc.contingency_factor for mcc in self.mccl))

        self.total_actual_contingency = int(sum(mcc.total_actual_contingency for mcc in self.mccl))

        # TX Size -> Transformer BoQ

@dataclass
class ClientMechanicalEquipmentList:
    """Class for Mechanical Equipment List."""

    mel: List[MechanicalEquipment] = field(default_factory=list)
    mcc_numbers: list = field(init=False)

    def __post_init__(self):

        self.mcc_numbers = []
        for me in self.mel:
            if me.mcc_number not in self.mcc_numbers:
                self.mcc_numbers.append(me.mcc_number)


def read_standards(standards):
    """ Create ProjectStandard object from a Project Standards excel file."""

    from openpyxl import load_workbook

    wb = load_workbook(filename=standards, data_only=True)
    ws = wb.active

    standards = ProjectStandard(
        ws["E5"].value,
        ws["E10"].value,
        ws["E11"].value,
        ws["E12"].value,
        ws["E15"].value,
        ws["E16"].value,
        ws["E17"].value,
        ws["E18"].value,
        ws["E21"].value,
        ws["E22"].value,
        ws["E23"].value,
        ws["E24"].value,
        ws["E25"].value,
        ws["E26"].value,
        0,
        0,
        ws["E37"].value,
        ws["E38"].value,
        ws["E39"].value,
        ws["C43"].value,
        ws["E43"].value,
        ws["C44"].value,
        ws["E44"].value,
        ws["D47"].value,
        ws["D49"].value,
    )

    return standards


def me_builder(row):
    """ Creats a Mechanical Equipment object"""

    me = MechanicalEquipment(
        str(row[0]),
        str(row[1]),
        str(row[2]),
        str(row[3]),
        str(row[4]),
        str(row[5]),
        float(row[6]),
        str(row[7]),
        float(row[8]),
        row[9],
        str(row[10]),
        str(row[11]),
    )

    return me


def client_mel_builder(rows):
    """ Creates a  ClientMechanicalEquipmentList from ME excel row data"""

    me_list = []

    for row in rows:
        me = me_builder(row)
        me_list.append(me)

    CMEL = ClientMechanicalEquipmentList(me_list)

    return CMEL


def mcc_builder(name, lighting_load, ups_load, fe_dist_load, mcc_me_list):
    """ Creates a MCC objecte"""

    mcc = MotorControlCenter(
        name,
        LightingEquipment(lighting_load),
        UPSEquipment(ups_load),
        FieldEquipment(fe_dist_load),
        mcc_me_list,
    )

    return mcc


def read_client_mel(mel):
    """Create a list of MechanicalEquipment objects from a MEL excel file."""

    from openpyxl import load_workbook

    wb = load_workbook(filename=mel, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(min_row=8, max_col=12, max_row=ws.max_row, values_only=True)

    return client_mel_builder(rows)


def eload(standards, mel):
    # Read in Project Standards excel file
    STANDARD = read_standards(standards)

    # Read in client Mechanical Equipment List
    MEL = read_client_mel(mel)

    # Initialise Motor Control Centers
    MCC = {}
    for number in MEL.mcc_numbers:

        # Filter equipment by mcc_number
        mcc_mel = []
        for me in MEL.mel:
            if me.mcc_number == number:
                mcc_mel.append(me)

        MCC[number] = mcc_builder(
            number,
            STANDARD.lighting_load,
            STANDARD.ups_load,
            STANDARD.fe_dist_load,
            mcc_mel,
        )
